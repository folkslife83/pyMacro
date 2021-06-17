import pyautogui
import time
import datetime
import NiconClicker
import Nheart
import os
import shutil
import sys
import webbrowser
import random
from sys import exit
from tkinter import *
import tkinter.messagebox
import tkinter.ttk as ttk
pyautogui.FAILSAFE = False
from checkBR import brOK, browser
from openpyxl import load_workbook
confidenceVal = 0.9

outClick = (10,0)
inClick = (10,200)
picClick = (10,600)
picCenter = (450,450)

def mult(num1, num2, num3,num4):    
    url = ("https://www.instagram.com/")
    webbrowser.open(url)  
    time.sleep(3)
    browser()
    time.sleep(1)
    for k in range (5):
        time.sleep(0.5)
        pyautogui.hotkey('ctrl','+')
    time.sleep(0.5)
    brOK()

    #nPages = num1   #창 몇개
    heartMax = num2 #최대하트수
    pgMax = num3    #최대 서치 오른쪽 넘기기
    load_wb = load_workbook("iList.xlsx", data_only=True)
    load_ws = load_wb['list'] #시트 이름으로 불러오기
    last_row = load_ws.active.max_row
    
    if num4 < 5:
        cell1st = 5
    elif num4> int(last_row):
        cell1st = int(last_row)
    else:
        cell1st = num4  #시작 셀
    
    if cell1st+num1 > int(last_row)+1:
        nPages = int(last_row) - cell1st + 1
    else:
        nPages = num1
    
    
    for i in range(nPages):
        id = load_ws.cell(cell1st+i,2).value   
        copy_string = "echo " + id + " |clip"
        os.system(copy_string)         #클립보드에 id올리기
        
        search = pyautogui.locateCenterOnScreen('images/USE/iSearch.png', confidence = confidenceVal)
        
   
        if(search):
            pyautogui.click(search)
            time.sleep(random.random()) #0과 1사이값                 

        else:
            continue

        pyautogui.hotkey('ctrl','v')
        time.sleep(1)
        pyautogui.typewrite(['enter'])
        time.sleep(2)
        pyautogui.typewrite(['enter'])
        time.sleep(2)
        pyautogui.click(picClick)
        time.sleep(1)
        pyautogui.typewrite(['pgdn'])
        time.sleep(2)
        time.sleep(random.random()) #0과 1사이값
        pyautogui.click(picClick)
        time.sleep(2)
        time.sleep(random.random()) #0과 1사이값
        for k in range(6):
            pyautogui.typewrite(['left'])
            time.sleep(random.random()) #0과 1사이값
        cnt =0
        time.sleep(2)
        for k in range(pgMax):
            pyautogui.moveTo(picCenter)
            time.sleep(0.5)
            pyautogui.scroll(-10000)
            time.sleep(0.5)
            Hempty = pyautogui.locateCenterOnScreen('images/USE/iHb.png', confidence = confidenceVal)
            #Hfull = pyautogui.locateCenterOnScreen('images/USE/iHr.png', confidence = confidenceVal)
            if Hempty:
                pyautogui.click(Hempty)
                cnt += 1
                time.sleep(random.random()) #0과 1사이값
                pyautogui.typewrite(['right'])
                
            else:
                pyautogui.typewrite(['right'])
        
        
        
        now = datetime.datetime.today()    
        end = ".txt"
        fname = now.strftime('%Y-%m-%d-%H%M%S_'+ str(cell1st+i) + end)
        f=open("iHgiven/"+fname, 'w',encoding="UTF8")
        f.write(id+"="+str(cnt) + "\n")
        f.close()
        time.sleep(random.random()) #0과 1사이값
        cnt = 0
        pyautogui.typewrite(['esc'])
        time.sleep(1) 


    time.sleep(2)
    exit()


   
def okClick():
    num1 = int(combx.get())
    num2 = int(combxH.get())
    num3 = int(combxP.get())
    num4 = int(combxS.get())
    
    mult(num1, num2, num3,num4)        



def okClickImage1(): #집컴 home
    pathHome = os.path.realpath('images/home') 
    path = os.path.realpath('images/USE')
    files = os.listdir(pathHome)
    for file in files:
        shutil.copyfile(pathHome+'/'+file, path+'/'+file)
    Msgbox1()

def okClickImage2(): #원장실 one
    pathOne = os.path.realpath('images/one') 
    path = os.path.realpath('images/USE')
    files = os.listdir(pathOne)
    for file in files:
        shutil.copyfile(pathOne+'/'+file, path+'/'+file)
    Msgbox1()

def okClickImage3(): #원장실sub
    pathOneSub = os.path.realpath('images/oneSub') 
    path = os.path.realpath('images/USE')
    files = os.listdir(pathOneSub)
    for file in files:
        shutil.copyfile(pathOneSub+'/'+file, path+'/'+file)
    Msgbox1()

def Msgbox1():
    tkinter.messagebox.showinfo("Inform","Image files've been copied.")

win = Tk()
win.geometry("500x500+1300+100")
win.resizable(True,True)
win.title("execute")

label1=Label(win, text="*Chrome 가장왼쪽모니터 전체화면")
label1.pack()
label1_1=Label(win, text="***인스타로그인/200%확대확인***")
label1_1.pack()

btn3 = Button(win, text = "***이미지수정***집pc", overrelief="solid", width=30, command=okClickImage1)
btn3.pack()
btn4 = Button(win, text = "***이미지수정***원장실", overrelief="solid", width=30, command=okClickImage2)
btn4.pack()
btn5 = Button(win, text = "***이미지수정***원장실sub", overrelief="solid", width=30, command=okClickImage3)
btn5.pack()


label55=Label(win, text="시작 셀 넘버")
label55.pack()

valS = [str(i) for i in (5, 10,100)]
combxS = ttk.Combobox(win, height=5, values=valS)
combxS.set(5)
combxS.pack()

label2=Label(win, text="작업대상 웹페이지 수 선택")
label2.pack()

val = [str(i) for i in (1,5,10,30, 50, 100, 200, 500)]
combx = ttk.Combobox(win, height=5, values=val)
combx.set(100)
combx.pack()



label3=Label(win, text="좋아요 최대 클릭수")
label3.pack()

valH = [str(i) for i in (1,3,5, 10)]
combxH = ttk.Combobox(win, height=5, values=valH)
combxH.set(3)
combxH.pack()

label4=Label(win, text="탐색 페이지 수")
label4.pack()

valP = [str(i) for i in (1,3,5, 10,30)]
combxP = ttk.Combobox(win, height=5, values=valP)
combxP.set(2)
combxP.pack()




btn1 = Button(win, text = "실행", background="cornflowerblue",overrelief="solid", width=15, command=okClick)
btn1.pack()

win.mainloop()



