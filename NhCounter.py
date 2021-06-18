import pyautogui
import time
import random
import webbrowser
from openpyxl import load_workbook, Workbook
from tkinter import *
import tkinter.ttk as ttk
from tkinter import Tk
from _tkinter import TclError
import time, ctypes
from checkBR import brOK
import glob
import os
pyautogui.FAILSAFE = False

# 네이버 로그인확인
# 왼쪽 모니터에 100% 배율상태확인
npages = 10 #최대 댓글이 매우 큰 숫자일 경우 높여준다.

left_win = (450,140)
right_winA = (1250,300)
right_winB = (1600,300)
def get_clipboard():
    root = Tk()
    root.withdraw()
    result = None
    try:
        result = root.clipboard_get()
    except TclError as e:
        print('empty clipboard')
    return result


def count():
    time.sleep(0.5)
    pyautogui.click(left_win)

    for i in range(npages):
        time.sleep(0.2)
        pyautogui.typewrite(['end'])
    
    time.sleep(0.5)
    pyautogui.hotkey('ctrl','a') 
    time.sleep(0.2)
    pyautogui.hotkey('ctrl','c') 
    time.sleep(0.2)

    pyautogui.click(right_winA)
    time.sleep(1)
    pyautogui.click(button='right')
    time.sleep(1)
    pyautogui.typewrite(['p'])
    time.sleep(2)
    pyautogui.click(right_winB)
    time.sleep(1)    
    pyautogui.hotkey('ctrl','q') 
    time.sleep(1)
    pyautogui.hotkey('ctrl','space') 
    time.sleep(1)
    pyautogui.hotkey('ctrl','c')     
    time.sleep(3)
    id = get_clipboard().split('\n')
    id =  [v for v in id if v] #빈문자열 제거
    time.sleep(0.5)
    pyautogui.hotkey('ctrl','a')
    time.sleep(0.5)
    pyautogui.typewrite(['del'])
    time.sleep(1)
    pyautogui.hotkey('ctrl','g')
    time.sleep(0.5)
    pyautogui.typewrite(['tab'])
    time.sleep(0.5)    
    pyautogui.typewrite(['tab'])
    time.sleep(0.5)
    pyautogui.typewrite(['enter'])
    time.sleep(0.5)            
    pyautogui.typewrite(['b'])
    time.sleep(0.5)
    pyautogui.typewrite(['enter'])
    time.sleep(0.5)        
    pyautogui.typewrite(['del'])
    time.sleep(1)    
    pyautogui.click(right_winB)
    time.sleep(1)
    pyautogui.click(left_win)
    time.sleep(1)
    pyautogui.hotkey('ctrl','w')
    time.sleep(1)

    lst = []
    for i in id[4:]:
        if i.find('blogId='):
            lst.append(i)
    lst2=[]
    for one in lst:
        lst2.append(one.split('=')[-1])
    return set(lst2)


    

def okClick():
    num1 = int(input_docNum1.get())
    num2 = int(input_docNum2.get())

    url_test = ("https://m.blog.naver.com/folkslife")
    webbrowser.open(url_test)  
    time.sleep(0.5)
    brOK()

    load_wb = load_workbook("nList.xlsx", data_only=True)
    load_ws = load_wb['name'] #시트 이름으로 불러오기
    last_row = load_ws.max_row
    lrow = int(last_row)

    if num1 < 1:                    #시작글번호
            docNum1 = 1
    elif num1> lrow-4:
        docNum1 = lrow-4
    else:
        docNum1 = num1  
    
    if  docNum1 > num2: 
        docNum2 = docNum1
    else:
        docNum2 = num2    

    for i in range (lrow-4):
        docNum = int(load_ws.cell(i+5,1).value) #글번호
               
        if docNum > docNum2 :
            continue
        if docNum < docNum1 :
            break

        num = (load_ws.cell(i+5,2).value).split('/')[-1]        
        url = "https://m.blog.naver.com/SympathyHistoryList.naver?blogId=folkslife&logNo="+ str(num) + "&categoryId=POST"
        webbrowser.open(url)
        time.sleep(1)    
        
        output = count()  
        fname = str(docNum) + ".txt"

        path = os.path.dirname(os.path.realpath(__file__))
        os.chdir(path)
        name = "C:\\Git\\pyMacro\\nHreceive\\"+ fname
        if os.path.exists(name):
            os.remove(name)


        f=open("C:\\Git\\pyMacro\\nHreceive\\"+fname, 'w',encoding="UTF8")
                
        for id in output:
            f.write(id + "\n")
        f.close()

        time.sleep(1)
        if docNum < 2 :
            break
    
    time.sleep(5)
    exit()

win = Tk()
win.geometry("320x150+1300+0")
win.resizable(True,True)
win.title("H counter")

label1=Label(win, text="*Chrome 가장왼쪽모니터 100%배율* ")
label1.pack()
label2=Label(win, text="***창열기전 네이버로그인//hLink열기!!!***")
label2.pack()

label3=Label(win, text="시작 글번호")
label3.pack()
input_docNum1 = Entry (win, width = 15)
input_docNum1.insert(0, 1)

label4=Label(win, text="마지막 글번호")
label4.pack()
input_docNum2 = Entry (win, width = 15)

btn1 = Button(win, text = "실행", background="cornflowerblue",overrelief="solid", width=15, command=okClick)
btn1.pack()

label3.place(x=30, y=50)
label4.place(x=180, y=50)
input_docNum1.place(x=30, y=70)
input_docNum2.place(x=180, y=70)
btn1.place(x=105,y=100)

win.mainloop()




