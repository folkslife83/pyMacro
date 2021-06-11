import pyautogui
import time
import datetime
import NiconClicker
import Nheart
import os
import shutil
import sys
import webbrowser
from sys import exit
from tkinter import *
import tkinter.ttk as ttk
pyautogui.FAILSAFE = False
from checkBR import brOK, browser
from openpyxl import load_workbook

def mult(num1, num2, num3,num4, simul):    
    url = ("https://m.blog.naver.com/folkslife")
    webbrowser.open(url)  
    time.sleep(0.5)
    browser()
    brOK()

    #nPages = num1   #창 몇개
    heartMax = num2 #최대하트수
    pgMax = num3    #최대 서치다운
    load_wb = load_workbook("blog_list.xlsm", data_only=True)
    load_ws = load_wb['list'] #시트 이름으로 불러오기
    last_row = load_wb.active.max_row
    
    if num4 < 5:
        cell1st = 5
    elif num4> int(last_row):
        cell1st = int(last_row)
    else:
        cell1st = num4  #시작 셀
    
    if cell1st+num1 > int(last_row)+1:
        nPages = int(last_row) - cell1st + 1
    else:
        nPages = num1
   

    aWork = 10      #한번에 작업할 양
    nWork = nPages//aWork
    extraWork = nPages%aWork
    
    lst=[]
    for i in range(nWork):
        for k in range(aWork):            
            url = load_ws.cell(cell1st+i*aWork+k,2).value
            webbrowser.open(url)     
        time.sleep(aWork) #웹페이지 1개당 1초 
        lst += work(aWork, heartMax, pgMax, simul)    
    for i in range(extraWork):
        url = load_ws.cell(cell1st+nWork*aWork+i,2).value
        webbrowser.open(url)
    time.sleep(extraWork)  #웹페이지 1개당 1초 
    lst += work(extraWork, heartMax, pgMax, simul)
    time.sleep(0.2)    
    

    cellLast = cell1st+nPages-1
    now = datetime.datetime.today()
    fname = now.strftime('%Y-%m-%d-%H%M%S_'+ str(cellLast) )
    
    if simul:
        fname = fname + "_SIMULATION.txt"
    else:
        fname = fname + ".txt"

    f=open("nHgiven/"+fname, 'w',encoding="UTF8")
    
    for i in range(len(lst)):
        f.write(lst[i] + "\n")
    f.close()


def work(nPages, heartMax, pgMax, simul):
    history = {}
    for i in range(nPages):       
        id=NiconClicker.icon()    
        time.sleep(1)
        if simul:
            history[id] = Nheart.heartSimul(heartMax,pgMax)
        else:
            history[id] = Nheart.heart(heartMax,pgMax)

        pyautogui.click((10,200))
        pyautogui.hotkey('ctrl','w')    
    lst = list(history.keys())
    lst.reverse()

    now = datetime.datetime.today()
    output = []
    test = not(simul)
    fname = now.strftime('%Y-%m-%d-%H%M%S'+ str(test) +'.txt')
    f=open("nHgiven/temp/"+fname, 'w',encoding="UTF8")
    for i in range(len(lst)):
        output.append(str(lst[i]) + "="+str(history[lst[i]])  )
        f.write(output[i] + "\n")
    f.close()

    return output

    
def winOK():
    return True
def winFail():
    sys.exit("기존 프로그램 탐색-종료후 재실행")
    
def okClick():
    num1 = int(combx.get())
    num2 = int(combxH.get())
    num3 = int(combxP.get())
    num4 = int(combxS.get())
    
    mult(num1, num2, num3,num4,False)        

def okClickSimul():
    num1 = int(combx.get())
    num2 = int(combxH.get())
    num3 = int(combxP.get())
    num4 = int(combxS.get())
    
    mult(num1, num2, num3, num4, True)    

def okClickImage1(): #집컴 home
    pathHome = os.path.realpath('images/home') 
    path = os.path.realpath('images/USE')
    files = os.listdir(pathHome)
    for file in files:
        shutil.copyfile(pathHome+'/'+file, path+'/'+file)

def okClickImage2(): #원장실 one
    pathOne = os.path.realpath('images/one') 
    path = os.path.realpath('images/USE')
    files = os.listdir(pathOne)
    for file in files:
        shutil.copyfile(pathOne+'/'+file, path+'/'+file)


def okClickImage3(): #원장실sub
    pathOneSub = os.path.realpath('images/oneSub') 
    path = os.path.realpath('images/USE')
    files = os.listdir(pathOneSub)
    for file in files:
        shutil.copyfile(pathOneSub+'/'+file, path+'/'+file)


win = Tk()
win.geometry("500x500+1300+100")
win.resizable(True,True)
win.title("execute")

label1=Label(win, text="*Chrome 가장왼쪽모니터 전체화면")
label1.pack()
label1_1=Label(win, text="***창열기전 네이버로그인필수***")
label1_1.pack()

btn3 = Button(win, text = "***이미지수정***집pc", overrelief="solid", width=30, command=okClickImage1)
btn3.pack()
btn4 = Button(win, text = "***이미지수정***원장실", overrelief="solid", width=30, command=okClickImage2)
btn4.pack()
btn5 = Button(win, text = "***이미지수정***원장실sub", overrelief="solid", width=30, command=okClickImage3)
btn5.pack()


label55=Label(win, text="시작 셀 넘버")
label55.pack()

valS = [str(i) for i in (5, 10,100)]
combxS = ttk.Combobox(win, height=5, values=valS)
combxS.set(5)
combxS.pack()

label2=Label(win, text="작업대상 웹페이지 수 선택")
label2.pack()

val = [str(i) for i in (1,5,10,30, 50, 100, 200, 500)]
combx = ttk.Combobox(win, height=5, values=val)
combx.set(100)
combx.pack()



label3=Label(win, text="좋아요 최대 클릭수")
label3.pack()

valH = [str(i) for i in (1,3,5, 10)]
combxH = ttk.Combobox(win, height=5, values=valH)
combxH.set(3)
combxH.pack()

label4=Label(win, text="탐색 페이지 수")
label4.pack()

valP = [str(i) for i in (1,3,5, 10,30)]
combxP = ttk.Combobox(win, height=5, values=valP)
combxP.set(2)
combxP.pack()




btn1 = Button(win, text = "실행", background="cornflowerblue",overrelief="solid", width=15, command=okClick)
btn1.pack()

btn2 = Button(win, text = "모의탐색", overrelief="solid", width=15, command=okClickSimul)
btn2.pack()

win.mainloop()



