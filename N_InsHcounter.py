import pyautogui
import time, ctypes
import datetime
import os
import shutil
import sys
from tkinter import *
import tkinter.ttk as ttk
from tkinter import Tk
from _tkinter import TclError
import webbrowser
import random
from openpyxl import load_workbook, Workbook
from checkBR import brOK, browser
import tkinter.messagebox
import os
pyautogui.FAILSAFE = False
coordU = (760,273)  #드래그 시작할 위치

#Chrome창 전체화면 왼쪽모니터, 다운로드바 없게

def get_clipboard():
    root = Tk()
    root.withdraw()
    result = None
    try:
        result = root.clipboard_get()
    except TclError as e:
        print('empty clipboard')
    return result

def opn(url):    
    webbrowser.open(url)
    time.sleep(1)
    

def okClick():
    num1 = int(input_docNum1.get())
    num2 = int(input_docNum2.get())
    url_test = ("https://www.instagram.com/highan_gs")
    webbrowser.open(url_test)  
    time.sleep(0.5)
    browser()
    for k in range (5):
        time.sleep(0.5)
        pyautogui.hotkey('ctrl','+')
    time.sleep(0.5)
    brOK()

    load_wb = load_workbook("insta_list.xlsm", data_only=True)
    load_ws = load_wb['name'] #시트 이름으로 불러오기
    last_row = load_wb.active.max_row
    lrow = int(last_row)

    if num1 < 1:                    #시작글번호
            docNum1 = 1
    elif num1> lrow-4:
        docNum1 = lrow-4
    else:
        docNum1 = num1  
    
    if  docNum1 > num2: 
        docNum2 = docNum1
    else:
        docNum2 = num2
    
    for i in range (lrow-4):        
        
        docNum = int(load_ws.cell(i+5,1).value) #글번호
               
        if docNum > docNum2 :
            continue
        if docNum < docNum1 :
            break
        url = load_ws.cell(i+5,2).value
        
        opn(url)     
        time.sleep(1)
        confidenceVal = 0.8 #0.7은 인식불가 
        # 좋아하는 첫id가 길면 글자가 밀리는 현상


        pyautogui.click((10,200))
        time.sleep(0.5)
        pyautogui.typewrite(['pagedown'])
        time.sleep(0.5)
        pyautogui.click((10,0))
        nLikeThis = pyautogui.locateCenterOnScreen('images/USE/iHm.png', confidence = confidenceVal)
      
        if(nLikeThis):                       
            pyautogui.click(nLikeThis)
            time.sleep(2)
        else:
            return

        
        clipTxt = "first"
        idList=[]
        
        while True:                        
            if clipTxt != str(get_clipboard()):
                clipTxt = str(get_clipboard())
                idList += clipTxt.split('\n')
                time.sleep(1)
                pyautogui.moveTo(coordU)
                pyautogui.drag(0,687,1,button='left')
                time.sleep(2)
                pyautogui.hotkey('ctrl','c')
                time.sleep(1)
                pyautogui.typewrite(['pagedown'])
                time.sleep(1)               
                pyautogui.click(coordU)

            else:
                break
        if len(idList) == 0:
            continue
    
        output = []
        for name in idList:   # 님의 프로필 사진  - 포함하는 원소만 추리기
            if '님의 프로필 사진' in name:  
                output.append(name[:-9])
            else:
                continue
        time.sleep(1)   
        pyautogui.hotkey('ctrl','w')
        
        ids = set(output)
        fname = str(docNum) + ".txt"
        f=open("iHreceive/"+fname, 'w',encoding="UTF8")
                
        for id in ids:
            f.write(id + "\n")
        f.close()

        time.sleep(1)
        
        if docNum < 2 :
            break
    

    
    time.sleep(1)
    for m in range (5):
        time.sleep(0.5)
    pyautogui.hotkey('ctrl','-')
    time.sleep(0.5)
    exit()


def okClickImage1(): #집컴 home
    pathHome = os.path.realpath('images/home') 
    path = os.path.realpath('images/USE')
    files = os.listdir(pathHome)
    for file in files:
        shutil.copyfile(pathHome+'/'+file, path+'/'+file)
    Msgbox1()

def okClickImage2(): #원장실 one
    pathOne = os.path.realpath('images/one') 
    path = os.path.realpath('images/USE')
    files = os.listdir(pathOne)
    for file in files:
        shutil.copyfile(pathOne+'/'+file, path+'/'+file)
    Msgbox1()

def okClickImage3(): #원장실sub
    pathOneSub = os.path.realpath('images/oneSub') 
    path = os.path.realpath('images/USE')
    files = os.listdir(pathOneSub)
    for file in files:
        shutil.copyfile(pathOneSub+'/'+file, path+'/'+file)
    Msgbox1()

def Msgbox1():
    tkinter.messagebox.showinfo("Inform","Image files've been copied.")

 
win = Tk()
win.geometry("350x300+1300+0")
win.resizable(True,True)
win.title("H counter")

label1=Label(win, text="*Chrome 가장왼쪽모니터 100%배율* ")
label1.pack()
label2=Label(win, text="***창열기전 인스타로그인!!!***")
label2.pack()
btn3 = Button(win, text = "***이미지수정***집pc", overrelief="solid", width=30, command=okClickImage1)
btn3.pack()
btn4 = Button(win, text = "***이미지수정***원장실", overrelief="solid", width=30, command=okClickImage2)
btn4.pack()
btn5 = Button(win, text = "***이미지수정***원장실sub", overrelief="solid", width=30, command=okClickImage3)
btn5.pack()

label3=Label(win, text="시작 글번호")
label3.pack()
input_docNum1 = Entry (win, width = 15)
input_docNum1.insert(0, 1)

label4=Label(win, text="마지막 글번호")
label4.pack()
input_docNum2 = Entry (win, width = 15)

btn1 = Button(win, text = "실행", background="cornflowerblue",overrelief="solid", width=15, command=okClick)
btn1.pack()

label3.place(x=30, y=200)
label4.place(x=180, y=200)
input_docNum1.place(x=30, y=220)
input_docNum2.place(x=180, y=220)
btn1.place(x=105,y=270)

win.mainloop()








